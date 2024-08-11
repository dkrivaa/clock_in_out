from openpyxl import load_workbook
from datetime import datetime, time, timedelta
import pandas as pd


# Getting list of employers
def customers():
    # Loading workbook
    book = load_workbook('my_file.xlsx')
    # Select sheet
    sheet = book['customers']
    # getting last row
    max_row = sheet.max_row

    # Making list of employers
    employer_list = []
    for cell in sheet[f'A2:A{max_row}']:
        employer_list.append(cell[0].value)

    return employer_list


def new_customer(name):
    # Loading workbook
    book = load_workbook('my_file.xlsx')
    # Select sheet
    sheet = book['customers']
    # getting last row
    max_row = sheet.max_row
    sheet[f'A{max_row + 1}'] = name
    book.save('my_file.xlsx')

    return "Customer Added"

def add_clock_in(data):
    # Loading workbook
    book = load_workbook('my_file.xlsx')
    # Select sheet
    sheet = book['hours']
    # Adding new entry
    sheet.append(data)
    # Save file
    book.save('my_file.xlsx')


def display_clok_out():
    # Making df
    df = pd.read_excel('my_file.xlsx', sheet_name='hours')
    df = df[pd.isna(df['יציאה'])]
    return df


def add_clock_out(t, row_index):
    # Loading workbook
    book = load_workbook('my_file.xlsx')
    # Select sheet
    sheet = book['hours']
    sheet[f'E{row_index + 2}'] = t

    # Calculating hours worked
    hours = (datetime.combine(datetime.today(), t) -
             datetime.combine(datetime.today(), sheet[f'D{row_index + 2}'].value))
    sheet[f'F{row_index + 2}'] = hours

    book.save('my_file.xlsx')

    return 'Added clock-out'





def manual_add(t, row_index):
    # Loading workbook
    book = load_workbook('my_file.xlsx')
    # Select sheet
    sheet = book['hours']
    sheet[f'E{row_index + 2}'] = t

    # Calculating hours worked
    hours = (datetime.combine(datetime.today(), t) -
             datetime.combine(datetime.today(), sheet[f'D{row_index + 2}'].value))
    sheet[f'F{row_index + 2}'] = hours

    book.save('my_file.xlsx')

    return 'Added clock-out'






